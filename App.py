# Requires the tkinter library
import tkinter as tk
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import filedialog

class TableApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("     Spearman's rank     ")
        self.configure(bg="lightblue")  # Set background color for the entire application
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self, text="     Number of Rows:     ").pack()
        self.rows_entry = tk.Entry(self)
        self.rows_entry.pack()

        tk.Label(self, text="  Number of Columns: 2   ").pack()

        self.create_button = tk.Button(self, text="      Create Table       ", command=self.create_table)
        self.create_button.pack()

        self.calculate_button = tk.Button(self, text="Calculate Spearman's Rank", command=self.sort_and_assign_ranks, state=tk.DISABLED)
        self.calculate_button.pack()

        self.export_button = tk.Button(self, text="Export as Spreadsheet", command=self.export_data, state=tk.DISABLED)
        self.export_button.pack()

        # Add a text widget to display the correlation coefficient
        self.rank_display = tk.Text(self, height=2, width=30)
        self.rank_display.pack()

    def create_table(self): 
        self.create_button.config(state=tk.DISABLED)
        num_input_columns = 2

        try:
            self.num_rows = int(self.rows_entry.get())
        except ValueError:
            quit()
            return

        # Define table_canvas within the create_table method
        table_canvas = tk.Canvas(self, bg="lightblue")  # Set background color for the table canvas
        table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        table_frame = tk.Frame(table_canvas, bg="lightblue")  # Set background color for the table frame
        table_canvas.create_window((0, 0), window=table_frame, anchor=tk.NW)

        self.input_entry_vars = []
        for i in range(self.num_rows):
            entry_row = []
            for j in range(num_input_columns):
                entry_var = tk.DoubleVar()
                entry = tk.Entry(table_frame, textvariable=entry_var, borderwidth=1, relief="solid", width=15)
                entry.grid(row=i, column=j, padx=1, pady=1, sticky="w")
                entry_row.append(entry_var)
            self.input_entry_vars.append(entry_row)

        self.input_entry_vars2 = []
        for i in range(self.num_rows):
            entry_row2 = []
            for j in range(4):
                entry_var2 = tk.DoubleVar()
                entry2 = tk.Entry(table_frame, textvariable=entry_var2, borderwidth=1, relief="solid", width=15)
                entry2.grid(row=i, column=j+num_input_columns, padx=1, pady=1, sticky="w")
                entry_row2.append(entry_var2)
            self.input_entry_vars2.append(entry_row2)

        table_frame.update_idletasks()
        table_canvas.config(scrollregion=table_canvas.bbox("all"))

        v_scrollbar = tk.Scrollbar(self, orient=tk.VERTICAL, command=table_canvas.yview)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        table_canvas.configure(yscrollcommand=v_scrollbar.set)
        self.calculate_button.config(state=tk.NORMAL)
        self.rank_display.insert(tk.END, "Insert data into columns      1 and 2") 

    def sort_and_assign_ranks(self):
        column1_data = []
        column2_data = []

        for i in range(self.num_rows):
            entry1_data = float(self.input_entry_vars[i][0].get())
            entry2_data = float(self.input_entry_vars[i][1].get())
            column1_data.append(entry1_data)
            column2_data.append(entry2_data)
        
        ranks1, ranks2 = self.assign_ranks(column1_data, column2_data)
        difference_squared, n = self.get_difference_squared(ranks1, ranks2)
        correlation = self.get_correlation(difference_squared, n)
        message = self.check_correlation_category(correlation)
        # Display correlation coefficient in the text widget
        self.update_rank_display(correlation, message)
        
        # Update the input_entry_vars2 with ranks and differences
        for i in range(self.num_rows):
            self.input_entry_vars2[i][0].set(ranks1[i])
            self.input_entry_vars2[i][1].set(ranks2[i])
            self.input_entry_vars2[i][2].set(ranks1[i] - ranks2[i])
            self.input_entry_vars2[i][3].set((ranks1[i] - ranks2[i]) ** 2)
        
        self.export_button.config(state=tk.NORMAL)
        return difference_squared

    def assign_ranks(self, list0, list1):
        ranks0 = []
        ranks1 = []

        # Create a list of tuples containing the original values and their indices
        values_indices0 = [(val, idx) for idx, val in enumerate(list0)]
        values_indices1 = [(val, idx) for idx, val in enumerate(list1)]

        # Sort the list of tuples by values
        values_indices0.sort(reverse=True)
        values_indices1.sort(reverse=True)

        # Assign ranks, handling ties by averaging ranks
        rank0 = 1
        rank1 = 1
        count0 = 1
        count1 = 1
        for i in range(1, len(values_indices0)):
            val0, idx0 = values_indices0[i]
            prev_val0, _ = values_indices0[i - 1]
            val1, idx1 = values_indices1[i]
            prev_val1, _ = values_indices1[i - 1]

            if val0 == prev_val0:
                count0 += 1
            else:
                avg_rank0 = (rank0 + (rank0 + count0 - 1)) / 2
                ranks0.extend([avg_rank0] * count0)
                rank0 += count0
                count0 = 1

            if val1 == prev_val1:
                count1 += 1
            else:
                avg_rank1 = (rank1 + (rank1 + count1 - 1)) / 2
                ranks1.extend([avg_rank1] * count1)
                rank1 += count1
                count1 = 1

        # Handling the last element
        avg_rank0 = (rank0 + (rank0 + count0 - 1)) / 2
        avg_rank1 = (rank1 + (rank1 + count1 - 1)) / 2
        ranks0.extend([avg_rank0] * count0)
        ranks1.extend([avg_rank1] * count1)

        # Restore the original order of ranks based on the indices
        original_order_ranks0 = [None] * len(ranks0)
        original_order_ranks1 = [None] * len(ranks1)
        for i, idx in enumerate(values_indices0):
            original_order_ranks0[idx[1]] = ranks0[i]
        for i, idx in enumerate(values_indices1):
            original_order_ranks1[idx[1]] = ranks1[i]

        return original_order_ranks0, original_order_ranks1

    def get_difference_squared(self, ranks1, ranks2):
        print("Data")
        print("-------------------------")
        print("ranks1")
        print(ranks1)
        print("ranks2")
        print(ranks2)
        total_sum = 0
        n = len(ranks1)
        for i in range(len(ranks1)):
            total_sum += (((ranks1[i]-ranks2[i])*(ranks1[i]-ranks2[i])))
        print("Total sum")
        print(total_sum)
        return total_sum, n

    def get_correlation(self, total_sum, n):
        return round( (1 - (6 * total_sum) / (n * (n ** 2 - 1))),4)

    def update_rank_display(self, correlation,message):
        self.rank_display.delete("1.0", tk.END)  # Clear previous content
        self.rank_display.insert(tk.END, f"Correlation: {correlation}\n{message}")
        #self.rank_display.config(state=tk.DISABLED)  # Disable the widget to prevent editing
        #enable line above to prevent box with correlation being edited after first edit (even by pressing calculate)
        
    def check_correlation_category(self,correlation):
        if correlation == 1:
            return "Perfect Positive Correlation"
        elif 0.7 <= correlation < 1:
            return "Strong Positive Correlation"
        elif 0.3 <= correlation < 0.7:
            return "Moderate Positive Correlation"
        elif 0.05 < correlation < 0.3:
            return "Weak Positive Correlation"
        elif correlation == 0:
            return "No Correlation"
        elif -0.05 <= correlation <= 0.05:
            return "No Correlation"
        elif -0.3 < correlation < -0.05:
            return "Weak Negative Correlation"
        elif -0.7 < correlation <= -0.3:
            return "Moderate Negative Correlation"
        elif -1 < correlation <= -0.7:
            return "Strong Negative Correlation"
        elif correlation == -1:  
            return "Perfect Negative Correlation"
        else:
            return "Invalid correlation coefficient"
        
    def export_data(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Spearman's Rank Data"
        
        data1 = [float(self.input_entry_vars[i][0].get()) for i in range(self.num_rows)]
        data2 = [float(self.input_entry_vars[i][1].get()) for i in range(self.num_rows)]
        rank1 = [self.input_entry_vars2[i][0].get() for i in range(self.num_rows)]
        rank2 = [self.input_entry_vars2[i][1].get() for i in range(self.num_rows)]
        d = [self.input_entry_vars2[i][2].get() for i in range(self.num_rows)]
        d_squared = [self.input_entry_vars2[i][3].get() for i in range(self.num_rows)]
        correlation = self.get_correlation(sum(d_squared), self.num_rows)

        header = ["Value: ","Data1", "Data2", "Rank1", "Rank2", "D", "D^2"]
        ws.append(header)
        for i in range(self.num_rows):
            ws.append([i, data1[i], data2[i], rank1[i], rank2[i], d[i], d_squared[i]])
        ws.append([])
        ws.append(["Correlation", correlation])

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        black_border = Border(left=Side(style='thin', color='000000'),
                          right=Side(style='thin', color='000000'),
                          top=Side(style='thin', color='000000'),
                          bottom=Side(style='thin', color='000000'))


        for cell in ws[1]:
            cell.fill = green_fill
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = black_border
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb.save(file_path)



        # For csvs, as an alternative to excel
        #file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        #if file_path:
        #    with open(file_path, mode='w', newline='') as file:
        #        writer = csv.writer(file)
        #        writer.writerow(["Value: ","Data1", "Data2", "Rank1", "Rank2", "D", "D^2"])
        #        for i in range(self.num_rows):
        #            writer.writerow([
        #            [i],data1[i], data2[i], rank1[i], rank2[i], d[i], d_squared[i]
        #            ])
        #        writer.writerow([])
        #        writer.writerow(["Correlation", correlation])



if __name__ == "__main__":
    app = TableApp()
    app.mainloop()
